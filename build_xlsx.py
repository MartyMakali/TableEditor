"""Baut eine Excel-Arbeitsmappe aus einer JSON-Beschreibung.

Aufruf:
    python build_xlsx.py                     # volle Mappe
    python build_xlsx.py --vorlage           # nur die Kopfzeilen
    python build_xlsx.py --vorlage --zeilen 50
"""
import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formula.translate import Translator
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Vorgaben, falls die Beschreibung keinen eigenen Format-Block mitbringt.
# Farben als ARGB (fuehrendes FF = deckend), sonst weicht der Alphakanal ab.
STANDARDFORMAT = {
    "schrift": "Arial",
    "groesse": 10,
    "kopf_farbe": "FF44546A",
    "kopf_schrift": "FFFFFFFF",
    "kopf_fett": True,
    "rahmen_farbe": "FFBFBFBF",
    "rahmen_stil": "thin",
}


def argb(wert, ersatz):
    """Nimmt '#44546A', '44546A' oder 'FF44546A' und liefert immer ARGB."""
    if not wert:
        return ersatz
    w = wert.lstrip("#").upper()
    return w if len(w) == 8 else "FF" + w


class Stil:
    """Buendelt die aus dem Format-Block abgeleiteten openpyxl-Objekte."""

    def __init__(self, fmt):
        f = {**STANDARDFORMAT, **(fmt or {})}
        self.font_daten = Font(name=f["schrift"], size=f["groesse"])
        self.font_kopf = Font(
            name=f["schrift"],
            size=f["groesse"],
            bold=bool(f["kopf_fett"]),
            color=argb(f["kopf_schrift"], STANDARDFORMAT["kopf_schrift"]),
        )
        kopf = argb(f["kopf_farbe"], STANDARDFORMAT["kopf_farbe"])
        self.fill_kopf = PatternFill(fill_type="solid", start_color=kopf, end_color=kopf)
        self.align_kopf = Alignment(horizontal="general", vertical="center", wrap_text=True)

        stil = f["rahmen_stil"]
        if stil in (None, "", "keiner"):
            self.rahmen = Border()
        else:
            seite = Side(style=stil, color=argb(f["rahmen_farbe"], STANDARDFORMAT["rahmen_farbe"]))
            self.rahmen = Border(left=seite, right=seite, top=seite, bottom=seite)


def formel_versetzen(formel, spalte, von, nach):
    """Schreibt eine Formel von Zeile `von` auf Zeile `nach` um."""
    ursprung = f"{get_column_letter(spalte)}{von}"
    ziel = f"{get_column_letter(spalte)}{nach}"
    return Translator(formel, origin=ursprung).translate_formula(ziel)


def blatt_bauen(ws, blatt, stil, vorlage=False, leerzeilen=0):
    spalten = blatt["spalten"]

    for c, spalte in enumerate(spalten, start=1):
        zelle = ws.cell(row=1, column=c, value=spalte["name"])
        zelle.font = stil.font_kopf
        zelle.fill = stil.fill_kopf
        zelle.alignment = stil.align_kopf
        zelle.border = stil.rahmen

    if vorlage:
        muster = blatt["zeilen"][0] if blatt["zeilen"] else []
        zeilen = [
            [
                formel_versetzen(wert, c, 2, i + 2)
                if isinstance(wert, str) and wert.startswith("=")
                else None
                for c, wert in enumerate(muster, start=1)
            ]
            for i in range(leerzeilen)
        ]
    else:
        zeilen = blatt["zeilen"]

    for r, zeile in enumerate(zeilen, start=2):
        for c, spalte in enumerate(spalten, start=1):
            wert = zeile[c - 1] if c - 1 < len(zeile) else None
            zelle = ws.cell(row=r, column=c, value=wert)
            zelle.font = stil.font_daten
            zelle.alignment = Alignment(
                horizontal=spalte.get("ausrichtung", "left"),
                vertical="top",
                wrap_text=bool(spalte.get("umbruch")),
            )
            zelle.border = stil.rahmen

    for c, spalte in enumerate(spalten, start=1):
        if spalte.get("breite"):
            ws.column_dimensions[get_column_letter(c)].width = spalte["breite"]

    if blatt.get("kopfhoehe"):
        ws.row_dimensions[1].height = blatt["kopfhoehe"]
    if not vorlage:
        for zeilennr, hoehe in (blatt.get("zeilenhoehen") or {}).items():
            ws.row_dimensions[int(zeilennr)].height = hoehe

    if blatt.get("kopf_fixieren"):
        ws.freeze_panes = "A2"
    if blatt.get("autofilter") and spalten:
        letzte = get_column_letter(len(spalten))
        ws.auto_filter.ref = f"A1:{letzte}{len(zeilen) + 1}"


def mappe_bauen(doc, vorlage=False, zeilen=0):
    stil = Stil(doc.get("format"))
    wb = Workbook()
    wb.remove(wb.active)
    for blatt in doc["blaetter"]:
        blatt_bauen(wb.create_sheet(title=blatt["name"]), blatt, stil, vorlage, zeilen)
    return wb


def main():
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("-i", "--input", default="data/anforderungskatalog.json")
    p.add_argument("-o", "--output", default="output/Anforderungskatalog.xlsx")
    p.add_argument("--vorlage", action="store_true", help="leere Mappe, nur die Kopfzeilen")
    p.add_argument(
        "--zeilen",
        type=int,
        default=0,
        metavar="N",
        help="mit --vorlage: N leere, vorformatierte Datenzeilen samt Formeln",
    )
    args = p.parse_args()

    doc = json.loads(Path(args.input).read_text(encoding="utf-8"))
    ziel = Path(args.output)
    ziel.parent.mkdir(parents=True, exist_ok=True)
    mappe_bauen(doc, vorlage=args.vorlage, zeilen=args.zeilen).save(ziel)

    print(f"geschrieben: {ziel}")
    for b in doc["blaetter"]:
        anzahl = args.zeilen if args.vorlage else len(b["zeilen"])
        print(f"  {b['name']}: {anzahl} Datenzeilen x {len(b['spalten'])} Spalten")


if __name__ == "__main__":
    main()
