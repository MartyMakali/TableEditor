"""Liest Tabellen aus Word-, Excel- und CSV-Dateien aus.

Jede gefundene Tabelle kommt als Wörterbuch zurück:
    {"quelle": "Bericht.docx", "bezeichnung": "Tabelle 2", "kopf": [...], "zeilen": [[...], ...]}
Die erste Zeile gilt jeweils als Kopfzeile.
"""
import csv
import io
import re
from pathlib import Path

import openpyxl

WORD = {".docx"}
EXCEL = {".xlsx", ".xlsm"}
TEXT = {".csv", ".tsv", ".txt"}
FORMATE = WORD | EXCEL | TEXT


class NichtLesbar(Exception):
    pass


def _sauber(wert):
    """Vereinheitlicht einen Zellwert: leere Zeichenketten werden zu None."""
    if wert is None:
        return None
    if isinstance(wert, str):
        wert = wert.replace("\r\n", "\n").strip()
        return wert or None
    return wert


GANZ = re.compile(r"^-?(0|[1-9]\d*)$")           # fuehrende Nullen bleiben Text ('007')
KOMMA = re.compile(r"^-?(0|[1-9]\d*)[.,]\d+$")   # 12,5 und 12.5


def _zahl_wenn_moeglich(wert):
    """Macht aus '12' die Zahl 12 — wie beim Einfuegen in Excel.

    Nur eindeutige Zahlen werden umgewandelt: Kennungen wie '007', 'Q90'
    oder Ausdruecke wie 'inf' bleiben Text.
    """
    if not isinstance(wert, str):
        return wert
    t = wert.strip()
    if GANZ.match(t):
        return int(t)
    if KOMMA.match(t):
        return float(t.replace(",", "."))
    return wert


def _zuschneiden(zeilen):
    """Entfernt komplett leere Zeilen und leere Spalten am rechten Rand."""
    zeilen = [z for z in zeilen if any(w is not None for w in z)]
    if not zeilen:
        return []
    breite = max(
        (max((i + 1 for i, w in enumerate(z) if w is not None), default=0) for z in zeilen),
        default=0,
    )
    return [list(z[:breite]) + [None] * (breite - len(z[:breite])) for z in zeilen]


def _als_tabelle(quelle, bezeichnung, zeilen, zahlen_erkennen=False):
    """Baut aus Rohzeilen eine Tabelle; die erste Zeile wird zur Kopfzeile.

    zahlen_erkennen gilt fuer Quellen ohne eigene Datentypen (Word, CSV);
    Excel bringt seine Typen schon mit.
    """
    zeilen = _zuschneiden(zeilen)
    if len(zeilen) < 1:
        return None
    kopf = [
        str(w).strip() if w is not None else f"Spalte {i + 1}"
        for i, w in enumerate(zeilen[0])
    ]
    daten = zeilen[1:]
    if zahlen_erkennen:
        daten = [[_zahl_wenn_moeglich(w) for w in z] for z in daten]
    return {
        "quelle": quelle,
        "bezeichnung": bezeichnung,
        "kopf": kopf,
        "zeilen": daten,
    }


def _aus_word(quelle, daten):
    import docx  # nur bei Bedarf laden, spart Startzeit

    dokument = docx.Document(io.BytesIO(daten))
    tabellen = []
    for nr, tabelle in enumerate(dokument.tables, start=1):
        zeilen = [[_sauber(z.text) for z in reihe.cells] for reihe in tabelle.rows]
        gebaut = _als_tabelle(quelle, f"Tabelle {nr}", zeilen, zahlen_erkennen=True)
        if gebaut:
            tabellen.append(gebaut)
    return tabellen


def _aus_excel(quelle, daten):
    # Zwei Durchgaenge: Werte bevorzugt, sonst die Formel als Text.
    werte = openpyxl.load_workbook(io.BytesIO(daten), data_only=True)
    formeln = openpyxl.load_workbook(io.BytesIO(daten))

    tabellen = []
    for ws in werte.worksheets:
        wf = formeln[ws.title]
        zeilen = [
            [
                _sauber(zelle.value if zelle.value is not None else wf.cell(zelle.row, zelle.column).value)
                for zelle in reihe
            ]
            for reihe in ws.iter_rows()
        ]
        gebaut = _als_tabelle(quelle, ws.title, zeilen)
        if gebaut:
            tabellen.append(gebaut)
    return tabellen


def _aus_text(quelle, daten):
    for kodierung in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            text = daten.decode(kodierung)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise NichtLesbar("Zeichensatz der Datei nicht erkannt")

    probe = text[:4096]
    try:
        trenner = csv.Sniffer().sniff(probe, delimiters=";,\t|").delimiter
    except csv.Error:
        trenner = "\t" if "\t" in probe else ";" if ";" in probe else ","

    zeilen = [[_sauber(w) for w in z] for z in csv.reader(io.StringIO(text), delimiter=trenner)]
    gebaut = _als_tabelle(quelle, Path(quelle).stem, zeilen, zahlen_erkennen=True)
    return [gebaut] if gebaut else []


def tabellen_aus_datei(dateiname, daten):
    """Liefert alle Tabellen einer hochgeladenen Datei."""
    endung = Path(dateiname).suffix.lower()
    if endung == ".doc":
        raise NichtLesbar(
            "Das alte .doc-Format wird nicht unterstützt — bitte in Word als .docx speichern."
        )
    if endung == ".xls":
        raise NichtLesbar(
            "Das alte .xls-Format wird nicht unterstützt — bitte in Excel als .xlsx speichern."
        )
    if endung not in FORMATE:
        raise NichtLesbar(f"Format {endung or '(ohne Endung)'} wird nicht unterstützt")

    if endung in WORD:
        tabellen = _aus_word(dateiname, daten)
    elif endung in EXCEL:
        tabellen = _aus_excel(dateiname, daten)
    else:
        tabellen = _aus_text(dateiname, daten)

    if not tabellen:
        raise NichtLesbar("In der Datei wurde keine Tabelle gefunden")
    return tabellen
